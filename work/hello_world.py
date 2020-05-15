#!/usr/bin/python

import os
import sys
script_path = os.path.dirname(__file__)
lib_path = script_path + "/../pyLib"
sys.path.append(lib_path)

from openpyxl import Workbook
from openpyxl import load_workbook

# write exel
workbook = Workbook()
sheet = workbook.active
sheet["A1"] = "hello"
sheet["B1"] = "world!"
workbook.save(filename="hello_world.xlsx")

# read exel
workbook = load_workbook(filename="hello_world.xlsx")
sheet = workbook.active
print (sheet["A1"].value)
print (sheet["B1"].value)
